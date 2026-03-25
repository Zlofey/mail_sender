import logging
from typing import Optional

from django.db import transaction, IntegrityError
from openpyxl import load_workbook

from app.models import MailLog
from app.tasks import send_email_task

logger = logging.getLogger(__name__)

# ожидаемые заголовки колонок из первой строки файла
EXPECTED_HEADERS = ["external_id", "user_id", "email", "subject", "message"]


class XLSXImporter:
    """Сервис для импорта рассылок из XLSX файла."""

    def __init__(self, file_path: str):
        """Инициализация импортера."""
        self.file_path = file_path
        # счётчики статистики — прямо в классе
        self.processed = 0  # всего обработано строк
        self.created = 0  # создано новых записей
        self.skipped = 0  # пропущено дубликатов
        self.errors = 0  # ошибки парсинга/валидации

    def run(self) -> "XLSXImporter":
        """Запускает процесс импорта. Возвращает self для доступа к статистике."""
        logger.info(f"Starting import from {self.file_path}")

        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            sheet = wb.active

            headers = self._parse_headers(sheet)
            if not headers:
                logger.error("Failed to parse headers")
                return self

            self._process_data_rows(sheet, headers)

            return self

        except FileNotFoundError:
            logger.error(f"File not found: {self.file_path}")
            raise
        except Exception as e:
            logger.exception(f"Unexpected error during import: {e}")
            raise
        finally:
            wb.close()

    def _parse_headers(self, sheet) -> Optional[list[str]]:
        """Парсит заголовки из первой строки файла."""
        try:
            first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = self._parse_header_cells(first_row)

            for expected in EXPECTED_HEADERS:
                if expected not in headers:
                    logger.error(f"Missing required column: {expected}")
                    return None

            return headers

        except StopIteration:
            return None

    def _parse_header_cells(self, first_row: tuple) -> list[str]:
        """
        Парсит сырые ячейки в список заголовков.

        - Пропускает пустые ячейки (None)
        - Конвертирует в строку
        - Убирает пробелы по краям
        - Приводит к нижнему регистру
        """
        headers = []
        for cell in first_row:
            if cell is None:
                continue
            header = str(cell).strip().lower()
            headers.append(header)
        return headers

    def _process_data_rows(self, sheet, headers: list[str]) -> None:
        """Обрабатывает строки данных (начиная со второй)."""
        for row in sheet.iter_rows(min_row=2, values_only=True):
            self.processed += 1  # 👈 прямо self.processed

            # пропуск пустых строк
            if not any(cell for cell in row):
                continue

            status = self._process_single_row(row, headers)

            if status == "created":
                self.created += 1  # 👈 прямо self.created
            elif status == "skipped":
                self.skipped += 1
            elif status == "error":
                self.errors += 1

    def _process_single_row(self, row: tuple, headers: list[str]) -> str:
        """Обработка одной строки файла."""
        try:
            # маппинг данных по заголовкам
            data = dict(zip(headers, row))

            # валидация обязательных полей
            external_id = data.get("external_id")
            email = data.get("email")

            if not external_id or not email:
                logger.warning(f"Missing required fields in row: {data}")
                return "error"

            # проверка на дубликат
            if MailLog.objects.filter(external_id=external_id).exists():
                logger.debug(f"Skipped duplicate external_id: {external_id}")
                return "skipped"

            # создание записи в транзакции
            with transaction.atomic():
                log_entry = MailLog.objects.create(
                    external_id=str(external_id),
                    user_id=str(data.get("user_id", "")),
                    email=email,
                    subject=str(data.get("subject", "")),
                    message=str(data.get("message", "")),
                    status="pending",
                )

                # отправка задачи в Celery
                send_email_task.delay(log_entry.id)

            return "created"

        except IntegrityError as e:
            # проверка на дубликат на уровне БД
            logger.warning(f"IntegrityError for external_id: {e}")
            return "skipped"
        except Exception as e:
            logger.exception(f"Error processing row: {e}")
            return "error"
